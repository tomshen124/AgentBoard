#!/usr/bin/env python3
"""
Read, write, format, and analyze Excel files (.xlsx).
Dependencies: openpyxl
"""

import argparse
import csv
import json
import os
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Missing dependency: openpyxl", file=sys.stderr)
    print("Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def auto_width(ws):
    """Auto-adjust column widths to fit content."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value) if cell.value is not None else ''
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)


def apply_header_style(ws):
    """Apply professional header styling to the first row."""
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def load_csv_data(path):
    """Load CSV file into list of lists (header + rows)."""
    with open(path, newline='', encoding='utf-8-sig') as f:
        return list(csv.reader(f))


def load_json_data(path):
    """Load JSON file into header + rows."""
    with open(path, encoding='utf-8') as f:
        data = json.load(f)
    if not isinstance(data, list) or not data:
        return []
    headers = list(data[0].keys())
    rows = [headers]
    for item in data:
        rows.append([str(item.get(h, '')) for h in headers])
    return rows


# ---------------------------------------------------------------------------
# Commands
# ---------------------------------------------------------------------------

def cmd_read(args):
    """Read and display an Excel file."""
    if not os.path.isfile(args.file):
        print(f"Error: File not found: {args.file}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(args.file, data_only=True)
    sheet_name = args.sheet or wb.active.title
    if sheet_name not in wb.sheetnames:
        print(f"Error: Sheet '{sheet_name}' not found. Available: {wb.sheetnames}", file=sys.stderr)
        sys.exit(1)

    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(c) if c is not None else '' for c in row])

    if not rows:
        print("(Empty sheet)")
        return

    fmt = args.format or 'table'

    if fmt == 'json':
        if len(rows) < 2:
            output = json.dumps(rows, indent=2, ensure_ascii=False)
        else:
            headers = rows[0]
            data = []
            for r in rows[1:]:
                data.append(dict(zip(headers, r)))
            output = json.dumps(data, indent=2, ensure_ascii=False)
    elif fmt == 'csv':
        import io
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerows(rows)
        output = buf.getvalue()
    else:  # table
        lines = []
        if rows:
            col_widths = [0] * len(rows[0])
            for r in rows:
                for i, c in enumerate(r):
                    if i < len(col_widths):
                        col_widths[i] = max(col_widths[i], len(c))
            for ri, r in enumerate(rows):
                line = ' | '.join(c.ljust(col_widths[i]) for i, c in enumerate(r) if i < len(col_widths))
                lines.append(line)
                if ri == 0:
                    lines.append('-+-'.join('-' * w for w in col_widths))
        output = '\n'.join(lines)

    if args.save:
        with open(args.save, 'w', encoding='utf-8') as f:
            f.write(output)
        print(f"Saved to: {args.save}", file=sys.stderr)
    else:
        print(output)


def cmd_create(args):
    """Create a new Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = args.sheet or 'Sheet1'

    # Load data
    rows = []
    if args.from_csv:
        rows = load_csv_data(args.from_csv)
    elif args.from_json:
        rows = load_json_data(args.from_json)

    # Add title row
    start_row = 1
    if args.title and rows:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(rows[0]))
        title_cell = ws.cell(row=1, column=1, value=args.title)
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")
        start_row = 2

    # Write data
    for ri, row in enumerate(rows):
        for ci, val in enumerate(row):
            ws.cell(row=start_row + ri, column=ci + 1, value=val)

    # Apply formatting
    if args.header_style and rows:
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for ci in range(len(rows[0])):
            cell = ws.cell(row=start_row, column=ci + 1)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    if args.auto_width:
        auto_width(ws)

    if args.freeze_header:
        ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

    wb.save(args.output)
    print(f"Created: {args.output} ({len(rows)} rows)")


def cmd_add_sheet(args):
    """Add a new sheet to an existing workbook."""
    if not os.path.isfile(args.file):
        print(f"Error: File not found: {args.file}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(args.file)
    sheet_name = args.sheet or 'NewSheet'
    ws = wb.create_sheet(title=sheet_name)

    rows = []
    if args.from_csv:
        rows = load_csv_data(args.from_csv)
    elif args.from_json:
        rows = load_json_data(args.from_json)

    for ri, row in enumerate(rows):
        for ci, val in enumerate(row):
            ws.cell(row=ri + 1, column=ci + 1, value=val)

    if rows:
        apply_header_style(ws)
        auto_width(ws)

    wb.save(args.file)
    print(f"Added sheet '{sheet_name}' to {args.file} ({len(rows)} rows)")


def cmd_format(args):
    """Apply formatting to an existing Excel file."""
    if not os.path.isfile(args.file):
        print(f"Error: File not found: {args.file}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(args.file)
    ws = wb.active

    if args.auto_width:
        auto_width(ws)
        print("Applied auto-width")

    if args.header_style:
        apply_header_style(ws)
        print("Applied header style")

    if args.freeze_header:
        ws.freeze_panes = 'A2'
        print("Froze header row")

    wb.save(args.file)
    print(f"Formatted: {args.file}")


def cmd_analyze(args):
    """Analyze an Excel file and show summary statistics."""
    if not os.path.isfile(args.file):
        print(f"Error: File not found: {args.file}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(args.file, data_only=True)

    print(f"File: {args.file}")
    print(f"Sheets: {wb.sheetnames}")
    print()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            print(f"  [{sheet_name}] (empty)")
            continue

        print(f"  [{sheet_name}]")
        print(f"    Rows: {len(rows) - 1} (excluding header)")
        print(f"    Columns: {len(rows[0])}")

        # Header names
        headers = [str(h) if h else f'Col{i+1}' for i, h in enumerate(rows[0])]
        print(f"    Headers: {', '.join(headers)}")

        # Basic stats for numeric columns
        if len(rows) > 1:
            for ci, header in enumerate(headers):
                values = []
                for r in rows[1:]:
                    if ci < len(r) and r[ci] is not None:
                        try:
                            values.append(float(r[ci]))
                        except (ValueError, TypeError):
                            pass
                if values:
                    avg = sum(values) / len(values)
                    print(f"    {header}: min={min(values):.2f}, max={max(values):.2f}, "
                          f"avg={avg:.2f}, count={len(values)}")
        print()


def cmd_formula(args):
    """Add formulas to an Excel file."""
    if not os.path.isfile(args.file):
        print(f"Error: File not found: {args.file}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(args.file)
    ws = wb.active

    if args.cell:
        ws[args.cell] = args.formula
        print(f"Set {args.cell} = {args.formula}")
    elif args.column:
        start = args.start_row or 2
        end = args.end_row or ws.max_row
        col = args.column
        count = 0
        for row_num in range(start, end + 1):
            formula = args.formula.replace('{row}', str(row_num))
            cell_ref = f"{col}{row_num}"
            ws[cell_ref] = formula
            count += 1
        print(f"Applied formula to {count} cells in column {col}")

    wb.save(args.file)
    print(f"Saved: {args.file}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description='Excel Processing Tool')
    sub = parser.add_subparsers(dest='command', required=True)

    # read
    p = sub.add_parser('read', help='Read an Excel file')
    p.add_argument('file', help='Input .xlsx file')
    p.add_argument('--sheet', help='Sheet name')
    p.add_argument('--format', choices=['table', 'json', 'csv'])
    p.add_argument('--save', help='Save output to file')

    # create
    p = sub.add_parser('create', help='Create a new Excel file')
    p.add_argument('output', help='Output .xlsx file')
    p.add_argument('--from-csv', help='Import from CSV')
    p.add_argument('--from-json', help='Import from JSON')
    p.add_argument('--sheet', help='Sheet name')
    p.add_argument('--title', help='Title row text')
    p.add_argument('--auto-width', action='store_true')
    p.add_argument('--header-style', action='store_true')
    p.add_argument('--freeze-header', action='store_true')

    # add-sheet
    p = sub.add_parser('add-sheet', help='Add sheet to existing workbook')
    p.add_argument('file', help='.xlsx file to modify')
    p.add_argument('--sheet', help='New sheet name')
    p.add_argument('--from-csv', help='Import from CSV')
    p.add_argument('--from-json', help='Import from JSON')

    # format
    p = sub.add_parser('format', help='Apply formatting')
    p.add_argument('file', help='.xlsx file to format')
    p.add_argument('--auto-width', action='store_true')
    p.add_argument('--header-style', action='store_true')
    p.add_argument('--freeze-header', action='store_true')

    # analyze
    p = sub.add_parser('analyze', help='Analyze Excel file')
    p.add_argument('file', help='Input .xlsx file')

    # formula
    p = sub.add_parser('formula', help='Add formulas')
    p.add_argument('file', help='.xlsx file to modify')
    p.add_argument('--cell', help='Target cell (e.g., E2)')
    p.add_argument('--column', help='Target column for batch formula')
    p.add_argument('--formula', required=True, help='Formula (use {row} for row number)')
    p.add_argument('--start-row', type=int, help='Start row for column formula')
    p.add_argument('--end-row', type=int, help='End row for column formula')

    args = parser.parse_args()
    commands = {
        'read': cmd_read,
        'create': cmd_create,
        'add-sheet': cmd_add_sheet,
        'format': cmd_format,
        'analyze': cmd_analyze,
        'formula': cmd_formula,
    }
    commands[args.command](args)


if __name__ == '__main__':
    main()
